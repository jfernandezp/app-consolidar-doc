import pandas as pd
import os
import shutil
import sys
from openpyxl import load_workbook
import streamlit as st

def copiar_carpetas_por_documento(excel_path, directorio_busqueda, output_dir, columna_estado='ESTADO'):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active  # Usar la hoja activa (puedes especificar una hoja si lo prefieres)
        columna_index = None
        
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == columna_estado:
                columna_index = col[0].column  # Columna donde colocar "Ok"
                break
            
        # Si la columna destino no existe, crearla
        if columna_index is None:
            columna_index = ws.max_column + 1
            ws.cell(row=1, column=columna_index, value=columna_estado)
        
        # Asegurarse de que la carpeta de salida exista
        os.makedirs(output_dir, exist_ok=True)
        
        # Obtener la lista de subcarpetas dentro del directorio de búsqueda
        subcarpetas = [f for f in os.listdir(directorio_busqueda) 
                    if os.path.isdir(os.path.join(directorio_busqueda, f))]
        
        all_paths = []
        for f in subcarpetas:
            fp = os.path.join(directorio_busqueda, f)
            all_paths.append(fp)
            
        # Iterar sobre los documentos
        for row in range(2, ws.max_row + 1):
            ESTADO = str(ws.cell(row=row, column=columna_index).value)
            if ESTADO == "OK":
                continue
            
            fecha = str(ws.cell(row=row, column=1).value)
            nro_doc = str(ws.cell(row=row, column=3).value)
            proveedor = str(ws.cell(row=row, column=6).value)
            nro_factura = str(ws.cell(row=row, column=7).value)
            #anio = str(ws.cell(row=row, column=12).value)
            doc_str = str(nro_doc).strip()  # Convertir a string y limpiar espacios
            
            encontrado = False
            try:
                for sub in all_paths:
                    for root, sub_dir, files in os.walk(sub):
                        if doc_str in sub_dir:
                            encontrado = True
                        else:
                            encontrado = False
                        if encontrado:
                            carpeta_origen = os.path.join(sub, doc_str)
                            mes = get_name_mes(fecha.split(".")[1])
                            anio = fecha.split(".")[2]
                            destino = os.path.join(output_dir, anio, mes, fecha ,proveedor ,nro_doc + " " + nro_factura)
                            try:
                                if not os.path.exists(destino):
                                    os.makedirs(destino)
                                shutil.copytree(carpeta_origen, destino, dirs_exist_ok=True)
                            except Exception as e:
                                 print("Error: " + str(e))
                            break
                    if encontrado:
                        break
            except Exception as e:
                ws.cell(row=row, column=columna_index, value='ERROR')
                wb.save(excel_path)
                continue
            
            if encontrado:
                ws.cell(row=row, column=columna_index, value='OK')
            else:
                ws.cell(row=row, column=columna_index, value='NO_ENCONTRADO')
                
            wb.save(excel_path)
        
        return "OK - FINALIZADO"
    except Exception as e:
        return str(e)
    
def get_name_mes(nMes):
    try:
        months = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio", 
            7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        return months.get(int(nMes), "ERROR")
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        file_name = exc_tb.tb_frame.f_code.co_filename
        line_number = exc_tb.tb_lineno
        return f"ERROR - En el archivo: {file_name}, línea: {line_number}, Mensaje: {str(e)}"

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)
    return path

# Streamlit Interface
st.title("Copiar Carpetas Según Documentos")

# Subir archivo Excel
excel_file = st.file_uploader("Sube el archivo Excel", type="xlsx")

# Selección de las 3 carpetas de búsqueda
st.session_state.dir_busqueda = st.text_input("Ruta del directorio de búsqueda", "")
st.session_state.output_dir = st.text_input("Ruta del directorio de salida", "")

# Botón para ejecutar el script
if st.button("Ejecutar"):
    if excel_file and st.session_state.dir_busqueda and st.session_state.output_dir:
        # result = copiar_carpetas_por_documento(
        #     excel_path=excel_file,
        #     directorio_busqueda=dir_busqueda,
        #     output_dir=output_dir
        # )
        # Mostrar ícono de "ejecución"
        with st.spinner("⚙️ Ejecutando proceso..."):
            resultado = copiar_carpetas_por_documento(
                excel_path=excel_file,
                directorio_busqueda=st.session_state.dir_busqueda,
                output_dir=st.session_state.output_dir
            )
        #st.success(resultado)
    else:
        st.error("Por favor, completa todos los campos.")
        
    

    # Mostrar ícono de "finalizado"
    st.success("✔️ Proceso completado")
    st.info(f"Resultado: {resultado}")
    st.info(f"Excel temporal usado: {excel_file}")
